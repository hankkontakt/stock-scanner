"""
app.py
======
Portfolio Manager + MarketScan REST API - öppnas i webbläsaren på http://localhost:5001

Starta med:  python app.py
Stäng med:   Ctrl+C i terminalen

Funktioner:
- Sök aktier med live autocomplete (yfinance Search)
- Lägg till/ta bort/redigera innehav
- Se live-priser och P&L för hela portföljen
- Sparar automatiskt till holdings.csv
- REST API v1: /api/v1/ (scoring, portfolio, markets, etc.)
- Webhook management endpoints
- Swagger UI: /api/v1/docs
"""

import base64
import csv
import io
import os
import threading
import webbrowser
from datetime import datetime
from pathlib import Path

import requests
from nacl import encoding, public

import yfinance as yf
from flask import Flask, g, jsonify, render_template, request, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core import config
from portfolio import watchlist as wl

app = Flask(__name__)
HOLDINGS_FILE = config.HOLDINGS_FILE


# ── Helpers ────────────────────────────────────────────────────────────────

def load_holdings() -> list:
    """Läs holdings.csv och returnera som lista av dicts."""
    path = Path(HOLDINGS_FILE)
    if not path.exists():
        return []
    try:
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = []
            for row in reader:
                rows.append({
                    "ticker":     row.get("ticker", "").strip().upper(),
                    "shares":     float(row.get("shares", 0)),
                    "cost_basis": float(row.get("cost_basis", 0)) if row.get("cost_basis") else None,
                })
            return rows
    except Exception as e:
        print(f"Fel vid läsning av holdings: {e}")
        return []


def save_holdings(holdings: list):
    """Spara lista av holdings tillbaka till CSV."""
    path = Path(HOLDINGS_FILE)
    try:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["ticker", "shares", "cost_basis"])
            writer.writeheader()
            for h in holdings:
                writer.writerow({
                    "ticker":     h["ticker"],
                    "shares":     h["shares"],
                    "cost_basis": h.get("cost_basis", ""),
                })
    except Exception as e:
        print(f"Fel vid sparning av holdings: {e}")


def _encrypt_secret(public_key_b64: str, value: str) -> str:
    """Encrypt a secret value with the repo's public key (GitHub requirement)."""
    key = public.PublicKey(public_key_b64.encode(), encoding.Base64Encoder)
    box = public.SealedBox(key)
    return base64.b64encode(box.encrypt(value.encode())).decode()


def sync_watchlist_to_github():
    """Push watchlist.json content to GitHub Actions secret WATCHLIST_JSON."""
    token = os.getenv("GITHUB_TOKEN")
    owner = os.getenv("GITHUB_OWNER")
    repo  = os.getenv("GITHUB_REPO")
    if not all([token, owner, repo]):
        return False, "GITHUB_TOKEN/OWNER/REPO saknas i .env"

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    base = f"https://api.github.com/repos/{owner}/{repo}"

    r = requests.get(f"{base}/actions/secrets/public-key", headers=headers, timeout=10)
    if r.status_code != 200:
        return False, f"Kunde inte hämta publik nyckel: {r.status_code}"
    pk_data = r.json()

    from pathlib import Path as _P
    wl_path = _P(config.WATCHLIST_FILE)
    if not wl_path.exists():
        return True, "Tom bevakningslista - inget att synka"
    json_content = wl_path.read_text(encoding="utf-8")
    encrypted    = _encrypt_secret(pk_data["key"], json_content)

    r = requests.put(
        f"{base}/actions/secrets/WATCHLIST_JSON",
        headers=headers,
        json={"encrypted_value": encrypted, "key_id": pk_data["key_id"]},
        timeout=10,
    )
    if r.status_code in (201, 204):
        return True, "OK"
    return False, f"GitHub API svarade {r.status_code}: {r.text}"


def sync_holdings_to_github():
    """Push holdings.csv content to GitHub Actions secret HOLDINGS_CSV."""
    token = os.getenv("GITHUB_TOKEN")
    owner = os.getenv("GITHUB_OWNER")
    repo  = os.getenv("GITHUB_REPO")
    if not all([token, owner, repo]):
        return False, "GITHUB_TOKEN/OWNER/REPO saknas i .env"

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    base = f"https://api.github.com/repos/{owner}/{repo}"

    # 1. Hämta repots publika nyckel
    r = requests.get(f"{base}/actions/secrets/public-key", headers=headers, timeout=10)
    if r.status_code != 200:
        return False, f"Kunde inte hämta publik nyckel: {r.status_code}"
    pk_data = r.json()

    # 2. Läs och kryptera holdings.csv
    csv_content = Path(HOLDINGS_FILE).read_text(encoding="utf-8")
    encrypted   = _encrypt_secret(pk_data["key"], csv_content)

    # 3. Uppdatera secreten
    r = requests.put(
        f"{base}/actions/secrets/HOLDINGS_CSV",
        headers=headers,
        json={"encrypted_value": encrypted, "key_id": pk_data["key_id"]},
        timeout=10,
    )
    if r.status_code in (201, 204):
        return True, "OK"
    return False, f"GitHub API svarade {r.status_code}: {r.text}"


def parse_avanza_csv(content: str) -> list:
    """Parse Avanza portfolio export (semicolon-separated, Swedish decimals)."""
    content = content.lstrip('﻿').strip()
    reader = csv.DictReader(io.StringIO(content), delimiter=';')

    def sv_float(s):
        if not s: return None
        s = str(s).strip().replace('\xa0', '').replace(' ', '').replace('%', '')
        s = s.replace('.', '').replace(',', '.')
        try: return float(s)
        except: return None

    rows = []
    for row in reader:
        name = (row.get('Beteckning') or row.get('Namn') or row.get('Värdepapper') or '').strip()
        if not name:
            continue
        typ = (row.get('Typ') or row.get('Typ av värdepapper') or '').strip().lower()
        if typ and typ not in ('aktie', 'etf', 'fond', ''):
            continue
        antal = sv_float(row.get('Antal') or row.get('Antal aktier') or '')
        if not antal or antal <= 0:
            continue
        avg = sv_float(row.get('Genomsnittligt anskaffningsvärde') or '')
        tot = sv_float(row.get('Anskaffningsvärde') or '')
        if avg and avg > 0:
            cost = round(avg, 2)
        elif tot and tot > 0:
            cost = round(tot / antal, 2)
        else:
            cost = None
        rows.append({'name': name, 'shares': antal, 'cost_basis': cost})
    return rows


def get_live_price(ticker: str) -> dict:
    """Hämta nuvarande kurs och namn från yfinance (snabb variant)."""
    try:
        t = yf.Ticker(ticker)
        info = t.fast_info
        return {
            "price":    round(float(info.last_price), 2) if info.last_price else None,
            "currency": info.currency or "USD",
            "name":     t.info.get("shortName") or ticker,
        }
    except Exception:
        return {"price": None, "currency": "?", "name": ticker}


# ── Routes ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/search")
def search():
    """Sök aktier via yfinance Search."""
    q = request.args.get("q", "").strip()
    if len(q) < 1:
        return jsonify([])
    try:
        results = yf.Search(q, max_results=8).quotes
        filtered = []
        for r in results:
            if r.get("quoteType") in ("EQUITY", "ETF"):
                filtered.append({
                    "ticker":   r.get("symbol", ""),
                    "name":     r.get("shortname") or r.get("longname") or "",
                    "exchange": r.get("exchange") or r.get("fullExchangeName") or "",
                    "type":     r.get("quoteType", ""),
                })
        return jsonify(filtered[:7])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/holdings", methods=["GET"])
def get_holdings():
    """Hämta alla innehav med live-priser och P&L."""
    holdings = load_holdings()
    enriched = []
    total_value = 0
    total_cost  = 0

    for h in holdings:
        live = get_live_price(h["ticker"])
        price  = live["price"]
        shares = h["shares"]
        cost   = h.get("cost_basis")

        market_value = round(shares * price, 2)  if price else None
        cost_total   = round(shares * cost,  2)  if cost  else None
        pnl          = round(market_value - cost_total, 2) if (market_value and cost_total) else None
        pnl_pct      = round(pnl / cost_total * 100, 2)   if (pnl is not None and cost_total) else None

        if market_value:
            total_value += market_value
        if cost_total:
            total_cost += cost_total

        enriched.append({
            "ticker":       h["ticker"],
            "name":         live["name"],
            "shares":       shares,
            "cost_basis":   cost,
            "currency":     live["currency"],
            "current_price":price,
            "market_value": market_value,
            "pnl":          pnl,
            "pnl_pct":      pnl_pct,
        })

    total_pnl     = round(total_value - total_cost, 2) if (total_value and total_cost) else None
    total_pnl_pct = round(total_pnl / total_cost * 100, 2) if (total_pnl is not None and total_cost) else None

    return jsonify({
        "holdings":     enriched,
        "total_value":  round(total_value, 2),
        "total_cost":   round(total_cost, 2),
        "total_pnl":    total_pnl,
        "total_pnl_pct":total_pnl_pct,
    })


@app.route("/api/holdings", methods=["POST"])
def add_holding():
    """Lägg till ett nytt innehav."""
    data = request.get_json()
    ticker     = data.get("ticker", "").strip().upper()
    shares     = float(data.get("shares", 0))
    cost_basis = float(data.get("cost_basis")) if data.get("cost_basis") else None

    if not ticker or shares <= 0:
        return jsonify({"error": "Ogiltigt ticker eller antal"}), 400

    holdings = load_holdings()

    # Kolla om tickern redan finns - uppdatera då istället
    for h in holdings:
        if h["ticker"] == ticker:
            h["shares"]     = shares
            h["cost_basis"] = cost_basis
            save_holdings(holdings)
            sync_holdings_to_github()
            return jsonify({"status": "updated", "ticker": ticker})

    holdings.append({"ticker": ticker, "shares": shares, "cost_basis": cost_basis})
    save_holdings(holdings)
    sync_holdings_to_github()
    return jsonify({"status": "added", "ticker": ticker})


@app.route("/api/holdings/<ticker>", methods=["DELETE"])
def remove_holding(ticker: str):
    """Ta bort ett innehav."""
    ticker   = ticker.upper()
    holdings = load_holdings()
    before   = len(holdings)
    holdings = [h for h in holdings if h["ticker"] != ticker]

    if len(holdings) == before:
        return jsonify({"error": "Ticker ej hittad"}), 404

    save_holdings(holdings)
    sync_holdings_to_github()
    return jsonify({"status": "removed", "ticker": ticker})


@app.route("/api/watchlist", methods=["GET"])
def get_watchlist():
    """Hämta alla aktier i bevakningslistan med live-priser."""
    items = wl.load_watchlist()
    enriched = []
    for item in items:
        live = get_live_price(item["ticker"])
        enriched.append({
            "ticker":   item["ticker"],
            "name":     live["name"] or item.get("name", item["ticker"]),
            "currency": live["currency"],
            "price":    live["price"],
            "added":    item.get("added", ""),
        })
    return jsonify(enriched)


@app.route("/api/watchlist", methods=["POST"])
def add_to_watchlist():
    """Lägg till en aktie i bevakningslistan."""
    data   = request.get_json()
    ticker = data.get("ticker", "").strip().upper()
    name   = data.get("name", "").strip()

    if not ticker:
        return jsonify({"error": "Ogiltigt ticker"}), 400

    is_new = wl.add_ticker(ticker, name)
    sync_watchlist_to_github()
    status = "added" if is_new else "already_exists"
    return jsonify({"status": status, "ticker": ticker})


@app.route("/api/watchlist/<ticker>", methods=["DELETE"])
def remove_from_watchlist(ticker: str):
    """Ta bort en aktie från bevakningslistan."""
    ticker = ticker.upper()
    removed = wl.remove_ticker(ticker)
    if not removed:
        return jsonify({"error": "Ticker ej hittad i bevakningslistan"}), 404
    sync_watchlist_to_github()
    return jsonify({"status": "removed", "ticker": ticker})


@app.route("/api/smallcap/universe", methods=["GET"])
def get_smallcap_universe():
    """Returnerar custom-tickers + basuniversumstatistik."""
    from smallcap.universe import load_custom, SMALLCAP_UNIVERSE, is_builtin
    custom = load_custom()
    return jsonify({
        "custom":      custom,
        "base_count":  len(SMALLCAP_UNIVERSE),
        "total_count": len(SMALLCAP_UNIVERSE) + sum(
            1 for c in custom if not is_builtin(c["ticker"])
        ),
    })


@app.route("/api/smallcap/universe", methods=["POST"])
def add_to_smallcap():
    """Lägger till en ticker i custom-listan."""
    data    = request.get_json()
    ticker  = data.get("ticker", "").strip().upper()
    name    = data.get("name", "").strip()
    segment = data.get("segment", "first_north")
    if not ticker:
        return jsonify({"error": "Ogiltigt ticker"}), 400
    from smallcap.universe import add_custom, is_builtin
    is_new  = add_custom(ticker, name, segment)
    builtin = is_builtin(ticker)
    return jsonify({
        "status":  "added" if is_new else "already_exists",
        "ticker":  ticker,
        "builtin": builtin,
    })


@app.route("/api/smallcap/universe/<ticker>", methods=["DELETE"])
def remove_from_smallcap(ticker: str):
    """Tar bort en ticker ur custom-listan."""
    ticker = ticker.upper()
    from smallcap.universe import remove_custom
    if not remove_custom(ticker):
        return jsonify({"error": "Ticker ej hittad i custom-listan"}), 404
    return jsonify({"status": "removed", "ticker": ticker})


@app.route("/api/import/avanza", methods=["POST"])
def import_avanza_preview():
    """Parse Avanza CSV and return rows with suggested ticker matches."""
    if 'file' not in request.files:
        return jsonify({"error": "Ingen fil"}), 400
    content = request.files['file'].read().decode('utf-8-sig')
    rows = parse_avanza_csv(content)
    if not rows:
        return jsonify({"error": "Kunde inte läsa filen - kontrollera att det är en Avanza-export"}), 400

    results = []
    for row in rows:
        ticker = ''
        suggestions = []
        try:
            hits = yf.Search(row['name'], max_results=5).quotes
            for r in hits:
                if r.get('quoteType') in ('EQUITY', 'ETF', 'MUTUALFUND'):
                    suggestions.append({
                        'ticker': r.get('symbol', ''),
                        'name':   r.get('shortname') or r.get('longname') or '',
                    })
            if suggestions:
                ticker = suggestions[0]['ticker']
        except Exception:
            pass
        results.append({
            'avanza_name': row['name'],
            'shares':      row['shares'],
            'cost_basis':  row['cost_basis'],
            'ticker':      ticker,
            'suggestions': suggestions[:3],
        })
    return jsonify(results)


@app.route("/api/import/avanza/confirm", methods=["POST"])
def import_avanza_confirm():
    """Save confirmed import rows to holdings.csv."""
    rows = request.get_json() or []
    if not rows:
        return jsonify({"error": "Inga rader"}), 400

    holdings = load_holdings()
    n_added = n_updated = 0
    for row in rows:
        ticker = row.get('ticker', '').strip().upper()
        shares = float(row.get('shares', 0))
        cost   = row.get('cost_basis')
        if not ticker or shares <= 0:
            continue
        found = False
        for h in holdings:
            if h['ticker'] == ticker:
                h['shares'] = shares
                h['cost_basis'] = cost
                n_updated += 1
                found = True
                break
        if not found:
            holdings.append({'ticker': ticker, 'shares': shares, 'cost_basis': cost})
            n_added += 1

    save_holdings(holdings)

    ok, msg = sync_holdings_to_github()
    return jsonify({"status": "ok", "added": n_added, "updated": n_updated, "github_sync": ok, "github_msg": msg})


@app.route("/api/holdings/export/excel")
def export_holdings_excel():
    """Exportera innehav till .xlsx med live-priser och P&L."""
    holdings = load_holdings()
    enriched = []
    for h in holdings:
        live = get_live_price(h["ticker"])
        price  = live["price"]
        shares = h["shares"]
        cost   = h.get("cost_basis")
        market_value = round(shares * price, 2)  if price else None
        cost_total   = round(shares * cost,  2)  if cost  else None
        pnl          = round(market_value - cost_total, 2) if (market_value and cost_total) else None
        pnl_pct      = round(pnl / cost_total * 100, 2)   if (pnl is not None and cost_total) else None
        enriched.append({
            "Ticker":         h["ticker"],
            "Namn":           live["name"],
            "Valuta":         live["currency"],
            "Antal":          shares,
            "Köpkurs (snitt)":cost if cost else 0,
            "Senaste kurs":   price if price else 0,
            "Marknadsvärde":  market_value if market_value else 0,
            "P&L":            pnl if pnl else 0,
            "P&L %":          pnl_pct if pnl_pct else 0,
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfölj"

    # Stildefinitioner
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0B1520", end_color="0B1520", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="182534"),
        right=Side(style="thin", color="182534"),
        top=Side(style="thin", color="182534"),
        bottom=Side(style="thin", color="182534"),
    )
    pnl_green = Font(color="00D4AA")
    pnl_red   = Font(color="FF4D6D")

    # Skriv headers
    headers = list(enriched[0].keys()) if enriched else []
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Skriv data
    for row_idx, item in enumerate(enriched, 2):
        for col_idx, key in enumerate(headers, 1):
            val = item[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = left_align
            else:
                cell.alignment = cell_align
            # Färgsätt P&L
            if key == "P&L" and val is not None:
                cell.font = pnl_green if val >= 0 else pnl_red
                cell.number_format = '#,##0.00'
            elif key == "P&L %" and val is not None:
                if val >= 0:
                    cell.font = pnl_green
                else:
                    cell.font = pnl_red
                cell.number_format = '0.00"%"'
            elif key in ("Marknadsvärde", "Köpkurs (snitt)", "Senaste kurs"):
                cell.number_format = '#,##0.00'
            elif key == "Antal":
                cell.number_format = '#,##0'

    # Auto-bredd på kolumner
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=len(enriched)+1):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"holdings_{__import__('datetime').datetime.now().strftime('%Y-%m-%d')}.xlsx",
    )


# ═══════════════════════════════════════════════════════════════════════════
# HEALTH CHECK ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/health")
def health_full():
    """Full health check — returnerar JSON med status för alla komponenter."""
    try:
        from core.monitoring.health import system_health_check
        info = system_health_check()
        http_status = 200 if info.get("status") == "healthy" else 503
        return jsonify(info), http_status
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/health/live")
def health_live():
    """Liveness — returnerar 200 om appen lever."""
    return jsonify({"status": "alive", "timestamp": __import__('datetime').datetime.now().isoformat()})


@app.route("/health/ready")
def health_ready():
    """Readiness — 200 om alla API-nycklar finns."""
    try:
        from core.monitoring.health import _check_api_keys
        keys = _check_api_keys()
        all_ok = all(v == "ok" for v in keys.values())
        return jsonify({"status": "ready" if all_ok else "not_ready", "api_keys": keys}), (200 if all_ok else 503)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/metrics")
def health_metrics():
    """Prometheus metrics (om installerat)."""
    try:
        from core.monitoring.metrics import MetricsCollector
        mc = MetricsCollector()
        return mc.get_prometheus_text(), 200, {"Content-Type": "text/plain; charset=utf-8"}
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/health/cache")
def health_cache():
    """Cache-status."""
    try:
        from core.cache_utils import CacheAnalytics
        return jsonify(CacheAnalytics.to_dict())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/health/data")
def health_data():
    """Data coverage."""
    try:
        reports_dir = Path(__file__).resolve().parent.parent / "reports"
        files = sorted(reports_dir.glob("scored_universe_*.parquet"), reverse=True)
        if files:
            df = pd.read_parquet(files[0])
            from core.monitoring.health import check_data_coverage
            coverage = check_data_coverage(df)
            return jsonify(coverage)
        return jsonify({"total_rows": 0, "coverage": "Ingen scored_universe-data tillgänglig"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/health/resources")
def health_resources():
    """Resource monitoring."""
    try:
        from core.monitoring.resources import track_disk_usage, get_data_growth_rate
        return jsonify({
            "disk": track_disk_usage(),
            "growth": get_data_growth_rate(),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════════
# MARKETSCAN REST API v1 - Blueprint, Middleware & Webhooks
# ══════════════════════════════════════════════════════════════════════════════════

# -- Import och registrera API v1 blueprint --
try:
    from web.api import api_v1
    from web.api.docs import register_docs_routes

    register_docs_routes(api_v1)
    app.register_blueprint(api_v1)
    _API_V1_REGISTERED = True
    print("  [API] REST API v1 registrerad: /api/v1/")
except Exception as e:
    _API_V1_REGISTERED = False
    print(f"  [API] Kunde inte registrera API v1: {e}")

# -- Import webhook manager for endpoints --
try:
    from core.webhooks.webhook_manager import WebhookManager
    _WEBHOOK_MANAGER = WebhookManager()
except Exception as e:
    _WEBHOOK_MANAGER = None
    print(f"  [Webhook] Kunde inte ladda webhook manager: {e}")

# -- CORS-stöd (try flask-cors first, fallback to manual headers) --
try:
    from flask_cors import CORS
    CORS(app, resources={r"/api/*": {"origins": "*"}})
    _CORS_ENABLED = True
except ImportError:
    _CORS_ENABLED = False


@app.after_request
def add_security_and_rate_limit_headers(response):
    """Lägg till CORS- och rate limit-headers på alla svar.

    - CORS headers (om flask-cors ej installerat)
    - Rate limit headers från API-auth
    - Cache-Control for API-svar
    """
    # Manuell CORS om flask-cors inte är installerad
    if not _CORS_ENABLED:
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type, X-API-Key, Authorization")
        response.headers.add("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")

    # Rate limit headers från require_api_key decorator
    remaining = getattr(g, "rate_limit_remaining", None)
    if remaining is not None:
        response.headers["X-RateLimit-Limit"] = str(getattr(g, "rate_limit_limit", 100))
        response.headers["X-RateLimit-Remaining"] = str(remaining)
        response.headers["X-RateLimit-Reset"] = str(getattr(g, "rate_limit_reset", 0))

    # Cache-Control for API-svar
    if response.content_type and "application/json" in response.content_type:
        response.headers.add("Cache-Control", "no-cache, no-store, must-revalidate")

    return response


@app.before_request
def log_api_requests():
    """Logga API-anrop (endast for /api/*)."""
    if request.path.startswith("/api/"):
        try:
            from core import logger
            logger.info(
                "API %s %s from %s",
                request.method,
                request.path,
                request.remote_addr or "unknown",
            )
        except Exception:
            pass


# ── Webhook Management Endpoints ─────────────────────────────────────────────────


@app.route("/api/webhooks", methods=["GET"])
def webhook_list():
    """Lista alla registrerade webhooks."""
    if _WEBHOOK_MANAGER is None:
        return jsonify({"error": "Webhook-manager ej tillganglig"}), 503
    webhooks = _WEBHOOK_MANAGER.list_webhooks()
    stats = _WEBHOOK_MANAGER.get_webhook_stats()
    return jsonify({"webhooks": webhooks, "stats": stats})


@app.route("/api/webhooks", methods=["POST"])
def webhook_register():
    """Registrera en ny webhook."""
    if _WEBHOOK_MANAGER is None:
        return jsonify({"error": "Webhook-manager ej tillganglig"}), 503

    data = request.get_json() or {}
    url = data.get("url", "").strip()
    events = data.get("events", [])
    secret = data.get("secret", "")

    if not url:
        return jsonify({"error": "URL kravs"}), 400
    if not events:
        return jsonify({"error": "Minst ett event kravs"}), 400

    webhook_id = _WEBHOOK_MANAGER.register_webhook(url, events, secret)
    if webhook_id is None:
        return jsonify({"error": "Ogiltig URL eller event"}), 400

    return jsonify({"status": "created", "id": webhook_id}), 201


@app.route("/api/webhooks/<webhook_id>", methods=["DELETE"])
def webhook_unregister(webhook_id: str):
    """Ta bort en webhook."""
    if _WEBHOOK_MANAGER is None:
        return jsonify({"error": "Webhook-manager ej tillganglig"}), 503

    if not _WEBHOOK_MANAGER.unregister_webhook(webhook_id):
        return jsonify({"error": "Webhook ej hittad"}), 404

    return jsonify({"status": "removed", "id": webhook_id})


@app.route("/api/webhooks/<webhook_id>/log", methods=["GET"])
def webhook_log(webhook_id: str):
    """Hamta leveranslogg for en webhook."""
    if _WEBHOOK_MANAGER is None:
        return jsonify({"error": "Webhook-manager ej tillganglig"}), 503

    log = _WEBHOOK_MANAGER.get_delivery_log(webhook_id)
    return jsonify({"deliveries": log})


@app.route("/api/webhooks/test", methods=["POST"])
def webhook_test():
    """Testa en webhook-URL med ett test-event."""
    if _WEBHOOK_MANAGER is None:
        return jsonify({"error": "Webhook-manager ej tillganglig"}), 503

    data = request.get_json() or {}
    url = data.get("url", "").strip()

    if not url:
        return jsonify({"error": "URL kravs"}), 400

    # Skapa en temporar webhook, trigga, ta bort
    webhook_id = _WEBHOOK_MANAGER.register_webhook(url, ["scan.completed"], "test")
    if not webhook_id:
        return jsonify({"error": "Ogiltig URL"}), 400

    webhook = _WEBHOOK_MANAGER.get_webhook(webhook_id)
    result = _WEBHOOK_MANAGER.deliver_webhook(
        webhook, "scan.completed",
        {"test": True, "message": "Detta ar ett testmeddelande fran MarketScan"},
    )
    _WEBHOOK_MANAGER.unregister_webhook(webhook_id)
    return jsonify({"result": result})


@app.route("/api/webhooks/stats", methods=["GET"])
def webhook_stats():
    """Hamta webhook-statistik."""
    if _WEBHOOK_MANAGER is None:
        return jsonify({"error": "Webhook-manager ej tillganglig"}), 503
    return jsonify(_WEBHOOK_MANAGER.get_webhook_stats())


# ── Start ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = 5001
    url  = f"http://localhost:{port}"
    print(f"\n🚀 Portfolio Manager startar...")
    print(f"   Web UI:     {url}")
    print(f"   API v1:     {url}/api/v1/")
    print(f"   API Docs:   {url}/api/v1/docs")
    print(f"   Swagger:    {url}/api/v1/swagger.json")
    print(f"   Health:     {url}/api/v1/health")
    print(f"   Stoppa: Ctrl+C\n")

    # Öppna webbläsaren automatiskt efter 1.5s
    def open_browser():
        import time; time.sleep(1.5)
        webbrowser.open(url)
    threading.Thread(target=open_browser, daemon=True).start()

    app.run(host="0.0.0.0", port=port, debug=False)
